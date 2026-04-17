#!/mnt/c/dev/postgres-mcp-venv-linux/bin/python

from __future__ import annotations

import argparse
import csv
import io
import json
import os
import re
import subprocess
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Iterable
from xml.etree import ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "generated"
NAME_OVERRIDE_FILE = BASE_DIR / "name_overrides.csv"
NICKNAME_FILE = BASE_DIR / "nicknames.csv"
DB_DSN = os.environ.get(
    "TEAM_CONTACTS_DSN",
    "postgresql://postgres:6523Tike@192.168.1.248:5432/postgres",
)
SEASON = os.environ.get("TEAM_CONTACTS_SEASON", "2026")

NS = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
IGNORED_CELL_VALUES = {"", "HALF"}

THEME = {
    "navy": "2F5496",
    "light": "D9E2F3",
    "gold": "C9A227",
    "ink": "1F1F1F",
    "white": "FFFFFF",
    "grey": "6B7280",
}

CAPTAIN_ATTACHMENT_MODE_DEFAULT = "all-in-section"
CAPTAIN_ATTACHMENT_MODE_ALIASES = {
    "1": "own-plus-reserves",
    "2": "own-next-plus-reserves",
    "3": "all-in-section",
    "own-plus-reserves": "own-plus-reserves",
    "ownplusreserves": "own-plus-reserves",
    "own+reserves": "own-plus-reserves",
    "own-reserves": "own-plus-reserves",
    "own-next-plus-reserves": "own-next-plus-reserves",
    "ownnextplusreserves": "own-next-plus-reserves",
    "own+next+reserves": "own-next-plus-reserves",
    "current": "own-next-plus-reserves",
    "all-in-section": "all-in-section",
    "allinsection": "all-in-section",
    "all": "all-in-section",
}
CAPTAIN_ATTACHMENT_MODE_LABELS = {
    "own-plus-reserves": "Own team plus reserves",
    "own-next-plus-reserves": "Own team, next team down, plus reserves",
    "all-in-section": "All team sheets in the same section",
}


@dataclass
class ContactCandidate:
    first_name: str
    last_name: str
    member_status: str
    email: str
    phone: str
    mobile: str
    share_contact_detail: str


@dataclass
class MemberCandidate:
    first_name: str
    last_name: str
    category: str
    email: str
    phone: str
    mobile: str


@dataclass
class SignupCandidate:
    first_name: str
    last_name: str
    category: str
    email: str


@dataclass
class JuniorMainContactCandidate:
    first_name: str
    last_name: str
    member_name: str
    main_contact_name: str
    main_contact_email: str
    main_contact_phone: str
    main_contact_mobile: str
    parent_share_contact_detail: str
    match_rule: str


@dataclass
class ReviewEntry:
    section: str
    source_name: str
    target_name: str
    reason: str
    team_name: str


@dataclass
class ContactDisplay:
    phone: str
    email: str
    consent: str
    no_consent_reason: str
    cross_family: bool = False
    cross_family_reason: str = ""


@dataclass
class CaptainEmailAttachment:
    team_name: str
    file_name: str
    kind: str


@dataclass
class CaptainEmailJob:
    source_doc: str
    section: str
    team_name: str
    next_team_name: str
    captain_name: str
    captain_email: str
    attachment_mode: str
    attachments: list[CaptainEmailAttachment]
    can_send: bool
    blocked_reason: str


def normalize_captain_attachment_mode(value: str) -> str:
    normalized = str(value or CAPTAIN_ATTACHMENT_MODE_DEFAULT).strip().casefold().replace("_", "-")
    if not normalized:
        return CAPTAIN_ATTACHMENT_MODE_DEFAULT
    try:
        return CAPTAIN_ATTACHMENT_MODE_ALIASES[normalized]
    except KeyError as exc:
        available = ", ".join(sorted(CAPTAIN_ATTACHMENT_MODE_LABELS.keys()))
        raise argparse.ArgumentTypeError(
            f"Unknown captain attachment mode '{value}'. Use one of: {available}, or 1/2/3."
        ) from exc


def parse_args() -> argparse.Namespace:
    mode_help = "\n".join([
        "Captain attachment modes:",
        "  1 / own-plus-reserves",
        "      Send the captain their own team sheet plus the reserves sheet.",
        "  2 / own-next-plus-reserves",
        "      Send the captain their own team sheet, the next team down, plus reserves.",
        "  3 / all-in-section",
        "      Send the captain every team sheet in their section.",
        "",
        "The same value can be provided via TEAM_CAPTAIN_ATTACHMENT_MODE.",
    ])
    parser = argparse.ArgumentParser(
        description="Generate Avondale team contact sheets and captain mailout manifests."
        ,
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=mode_help,
    )
    parser.add_argument(
        "--captain-attachment-mode",
        type=normalize_captain_attachment_mode,
        default=normalize_captain_attachment_mode(
            os.environ.get("TEAM_CAPTAIN_ATTACHMENT_MODE", CAPTAIN_ATTACHMENT_MODE_DEFAULT)
        ),
        help=(
            "Captain attachment mode: "
            "own-plus-reserves (1), own-next-plus-reserves (2), or all-in-section (3, default)."
        ),
    )
    return parser.parse_args()

def run_query(sql: str) -> list[dict[str, str]]:
    result = subprocess.run(
        ["psql", DB_DSN, "--csv", "-c", sql],
        check=True,
        capture_output=True,
        text=True,
    )
    return list(csv.DictReader(io.StringIO(result.stdout)))


def normalize_name(value: str) -> str:
    cleaned = " ".join((value or "").replace("\xa0", " ").split())
    cleaned = cleaned.replace("–", "-").replace("—", "-")
    cleaned = re.sub(r"[^A-Za-z0-9]+", "", cleaned)
    return cleaned.casefold()


def load_name_overrides() -> dict[str, set[str]]:
    overrides: dict[str, set[str]] = defaultdict(set)
    if not NAME_OVERRIDE_FILE.exists():
        return overrides
    with NAME_OVERRIDE_FILE.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            source = normalize_name(row.get("source", ""))
            target = normalize_name(row.get("target", ""))
            if source and target:
                overrides[source].add(target)
    return overrides


def load_nicknames() -> dict[str, set[str]]:
    nicknames: dict[str, set[str]] = defaultdict(set)
    if not NICKNAME_FILE.exists():
        return nicknames
    with NICKNAME_FILE.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            source = normalize_name(row.get("source", ""))
            target = normalize_name(row.get("target", ""))
            if source and target:
                nicknames[source].add(target)
    return nicknames


EXACT_NAME_OVERRIDES = load_name_overrides()
NICKNAME_OVERRIDES = load_nicknames()


def clean_name(value: str) -> str:
    value = " ".join((value or "").replace("\xa0", " ").split())
    value = value.replace("–", "-").replace("—", "-")
    value = re.sub(r"\s*[©]\s*$", "", value)
    value = re.sub(r"\s*\(\s*c\s*\)\s*$", "", value, flags=re.I)
    value = re.sub(r"\s+c\s*$", "", value, flags=re.I)
    value = re.sub(r"\s+capt\.?\s*$", "", value, flags=re.I)
    value = re.sub(r"[.,;:]+$", "", value)
    return " ".join(value.split())


def is_captain(value: str) -> bool:
    value = value or ""
    return bool(
        re.search(r"(©|\(\s*c\s*\)|\bc\b|\bcapt\.?\b)\s*$", value.strip(), flags=re.I)
    )


def clean_phone(value: str) -> str:
    value = (value or "").strip()
    value = value.strip("[]")
    value = re.sub(r"\s+", " ", value)
    return value


def truthy_share_contact_detail(value: str) -> bool:
    normalized = re.sub(r"[^a-z0-9]+", "", (value or "").casefold())
    return normalized in {"yes", "y", "true", "1"}


def normalize_email_local(value: str) -> str:
    local = (value or "").split("@", 1)[0]
    return re.sub(r"[^a-z0-9]+", "", local.casefold())


def first_name_key(value: str) -> str:
    parts = value.split()
    return parts[0].casefold() if parts else ""


def split_person_name(value: str) -> tuple[str, str]:
    parts = value.split()
    if len(parts) <= 1:
        return value, ""
    return parts[0], " ".join(parts[1:])


def candidate_key(candidate: object) -> str:
    return normalize_name(f"{candidate.first_name} {candidate.last_name}")


def candidate_full_name(candidate: object) -> str:
    return f"{candidate.first_name} {candidate.last_name}".strip()


def choose_display_name(existing: str, new: str) -> str:
    if not existing:
        return new
    if existing.isupper() and not new.isupper():
        return new
    return existing


def doc_title_from_filename(path: Path) -> str:
    title = path.stem.replace("SUMMER", "Summer")
    title = title.replace("SQUADS", "Squads")
    title = re.sub(r"\s+", " ", title)
    return title.strip()


def sheet_title(header: str) -> str:
    words = [w.capitalize() for w in header.split()]
    value = " ".join(words)
    return value[:31]


def split_cell_entries(paragraphs: Iterable[str]) -> list[str]:
    raw_entries: list[str] = []
    for paragraph in paragraphs:
        parts = re.split(r"\.\s+(?=[A-Z])", paragraph.strip())
        for part in parts:
            cleaned = part.strip(" .")
            if cleaned:
                raw_entries.append(cleaned)

    entries: list[str] = []
    single_word_buffer: list[str] = []
    for entry in raw_entries:
        if len(entry.split()) == 1:
            single_word_buffer.append(entry)
            continue
        if single_word_buffer:
            entries.append(" ".join(single_word_buffer))
            single_word_buffer = []
        entries.append(entry)

    if single_word_buffer:
        entries.append(" ".join(single_word_buffer))

    return entries


def parse_docx_teams(path: Path) -> dict[str, list[dict[str, object]]]:
    with zipfile.ZipFile(path) as archive:
        root = ET.fromstring(archive.read("word/document.xml"))

    table = root.find(".//w:tbl", NS)
    if table is None:
        raise RuntimeError(f"No table found in {path.name}")

    row_cells: list[list[list[str]]] = []
    for tr in table.findall("./w:tr", NS):
        row: list[list[str]] = []
        for tc in tr.findall("./w:tc", NS):
            paragraphs: list[str] = []
            for p in tc.findall("./w:p", NS):
                text = "".join((t.text or "") for t in p.findall(".//w:t", NS))
                text = " ".join(text.replace("\xa0", " ").split())
                if text:
                    paragraphs.append(text)
            row.append(paragraphs)
        row_cells.append(row)

    if not row_cells:
        return {}

    headers = [" ".join(cell).strip() for cell in row_cells[0][1:]]
    teams: dict[str, list[dict[str, object]]] = {header: [] for header in headers if header}

    for row in row_cells[1:]:
        for index, header in enumerate(headers, start=1):
            if not header or index >= len(row):
                continue
            for paragraph_text in split_cell_entries(row[index]):
                cleaned = clean_name(paragraph_text)
                if cleaned in IGNORED_CELL_VALUES:
                    continue
                teams[header].append(
                    {
                        "name": cleaned,
                        "captain": is_captain(paragraph_text),
                    }
                )
    return teams


def load_lookup_data() -> tuple[
    dict[str, list[MemberCandidate]],
    dict[str, list[ContactCandidate]],
    dict[str, list[SignupCandidate]],
    dict[str, list[JuniorMainContactCandidate]],
]:
    members_sql = f"""
        select
          trim(m."First name") as first_name,
          trim(m."Last name") as last_name,
          mp.category as category,
          coalesce(nullif(trim(m."Email address"), ''), '') as email_address,
          coalesce(nullif(trim(m."Phone number"), ''), '') as phone_number,
          coalesce(nullif(trim(m."Mobile number"), ''), '') as mobile_number
        from public.raw_members m
        join public.membership_packages mp
          on mp.name = m."Membership"
        where coalesce(m.is_current, true) = true
          and mp.season = '{SEASON}'
          and coalesce(m."Status", '') = 'Active'
          and mp.category not in ('Social', 'Pavilion Key')
        order by 1, 2;
    """
    contacts_sql = """
        select
          c.first_name,
          c.last_name,
          coalesce(c.member_status, '') as member_status,
          c.norm_name,
          coalesce(c.email_address, '') as email_address,
          coalesce(c.phone_number, '') as phone_number,
          coalesce(c.mobile_number, '') as mobile_number,
          coalesce(nullif(trim(rc."Share Contact Detail"), ''), '') as share_contact_detail
        from public.vw_best_current_contacts c
        left join public.raw_contacts rc
          on rc.id = c.contact_id
        order by norm_name, contact_id;
    """
    signups_sql = f"""
        select
          trim(coalesce(ms."First name", split_part(ms.member, ' ', 1))) as first_name,
          trim(coalesce(ms."Last name", regexp_replace(ms.member, '^[^ ]+\\s*', ''))) as last_name,
          mp.category as category,
          coalesce(nullif(trim(ms.email_address), ''), '') as email_address
        from public.member_signups ms
        join public.membership_packages mp
          on mp.name = ms.product
        where mp.season = '{SEASON}'
          and coalesce(ms.status, '') = 'Complete'
          and mp.category not in ('Social', 'Pavilion Key')
        order by 1, 2;
    """
    junior_main_contacts_sql = """
        select
          member_name,
          main_contact_name,
          coalesce(main_contact_email, '') as main_contact_email,
          coalesce(main_contact_phone, '') as main_contact_phone,
          coalesce(main_contact_mobile, '') as main_contact_mobile,
          coalesce(nullif(trim(rc."Share Contact Detail"), ''), '') as parent_share_contact_detail,
          match_rule
        from public.vw_junior_main_contacts
        left join public.raw_contacts rc
          on rc.id = resolved_contact_id
        where match_confidence = 'high'
        order by member_name;
    """

    members_by_name: dict[str, list[MemberCandidate]] = defaultdict(list)
    contacts_by_name: dict[str, list[ContactCandidate]] = defaultdict(list)
    signups_by_name: dict[str, list[SignupCandidate]] = defaultdict(list)
    junior_main_contacts_by_name: dict[str, list[JuniorMainContactCandidate]] = defaultdict(list)

    for row in run_query(members_sql):
        members_by_name[normalize_name(f'{row["first_name"]} {row["last_name"]}')].append(
            MemberCandidate(
                first_name=row["first_name"],
                last_name=row["last_name"],
                category=row["category"],
                email=row["email_address"],
                phone=clean_phone(row["phone_number"]),
                mobile=clean_phone(row["mobile_number"]),
            )
        )

    for row in run_query(contacts_sql):
        contacts_by_name[normalize_name(f'{row["first_name"]} {row["last_name"]}')].append(
            ContactCandidate(
                first_name=row["first_name"],
                last_name=row["last_name"],
                member_status=row["member_status"],
                email=row["email_address"],
                phone=clean_phone(row["phone_number"]),
                mobile=clean_phone(row["mobile_number"]),
                share_contact_detail=row["share_contact_detail"],
            )
        )

    for row in run_query(signups_sql):
        signups_by_name[normalize_name(f'{row["first_name"]} {row["last_name"]}')].append(
            SignupCandidate(
                first_name=row["first_name"],
                last_name=row["last_name"],
                category=row["category"],
                email=row["email_address"],
            )
        )

    for row in run_query(junior_main_contacts_sql):
        first_name, last_name = split_person_name(row["member_name"])
        junior_main_contacts_by_name[normalize_name(row["member_name"])].append(
            JuniorMainContactCandidate(
                first_name=first_name,
                last_name=last_name,
                member_name=row["member_name"],
                main_contact_name=row["main_contact_name"],
                main_contact_email=row["main_contact_email"],
                main_contact_phone=clean_phone(row["main_contact_phone"]),
                main_contact_mobile=clean_phone(row["main_contact_mobile"]),
                parent_share_contact_detail=row["parent_share_contact_detail"],
                match_rule=row["match_rule"],
            )
        )

    return members_by_name, contacts_by_name, signups_by_name, junior_main_contacts_by_name


def choose_contact_detail(
    member_candidates: list[MemberCandidate],
    contact_candidates: list[ContactCandidate],
    signup_candidates: list[SignupCandidate],
    junior_main_contact_candidates: list[JuniorMainContactCandidate],
    all_contact_candidates: list[ContactCandidate] | None = None,
) -> ContactDisplay:
    self_contact = contact_candidates[0] if len(contact_candidates) == 1 else None
    main_contact = junior_main_contact_candidates[0] if len(junior_main_contact_candidates) == 1 else None
    has_self_consent = self_contact is not None and truthy_share_contact_detail(
        self_contact.share_contact_detail
    )
    has_parent_consent = main_contact is not None and truthy_share_contact_detail(
        main_contact.parent_share_contact_detail
    )

    self_phone = (self_contact.mobile or self_contact.phone) if self_contact else ""
    self_email = self_contact.email if self_contact else ""
    parent_phone = (
        main_contact.main_contact_mobile or main_contact.main_contact_phone
        if main_contact
        else ""
    )
    parent_email = main_contact.main_contact_email if main_contact else ""

    phone = ""
    email = ""
    cross_family = False
    cross_family_reason = ""

    # Collect contacts from all member candidates
    all_member_phones = []
    all_member_emails = []
    if member_candidates:
        for member in member_candidates:
            if member.mobile:
                all_member_phones.append(member.mobile)
            if member.phone:
                all_member_phones.append(member.phone)
            if member.email:
                all_member_emails.append(member.email)

    # If no direct contact or it has no useful details, check for cross-family contacts
    if (not self_contact or (self_contact and not self_contact.mobile and not self_contact.phone and not self_contact.email)) and member_candidates:
        surname = normalize_name(member_candidates[0].last_name)
        # Use all_contact_candidates if provided, otherwise use disambiguated list
        family_contacts = all_contact_candidates if all_contact_candidates is not None else contact_candidates
        # Collect contact details from all contacts with same surname (excluding self_contact)
        for contact in family_contacts:
            if normalize_name(contact.last_name) == surname and contact is not self_contact:
                if contact.mobile or contact.phone:
                    all_member_phones.append(contact.mobile or contact.phone)
                    cross_family = True
                if contact.email:
                    all_member_emails.append(contact.email)
                    cross_family = True
        if cross_family:
            cross_family_reason = f"contact details from family member with same surname"

    if len(member_candidates) == 1:
        member = member_candidates[0]
        if self_contact is None:
            phone = phone or member.mobile or member.phone
            email = email or member.email

    if self_contact is not None:
        phone = phone or self_phone
        email = email or self_email

    if len(signup_candidates) == 1:
        signup = signup_candidates[0]
        if self_contact is None:
            email = email or signup.email

    # Add any cross-family contacts if still no contact info
    if not phone and all_member_phones:
        phone = all_member_phones[0]
    if not email and all_member_emails:
        email = all_member_emails[0]

    self_phone = self_phone or phone
    self_email = self_email or email

    self_visible = bool(self_phone or self_email)
    parent_visible = bool(parent_phone or parent_email)

    def join_lines(lines: list[str]) -> str:
        return "\n".join(lines)

    def consent_text(allowed: bool) -> str:
        return "Yes" if allowed else "No"

    no_consent_parts: list[str] = []
    if self_visible and not has_self_consent:
        no_consent_parts.append("self")
    if parent_visible and not has_parent_consent:
        no_consent_parts.append("parent")

    if self_visible and parent_visible:
        phone_lines: list[str] = []
        email_lines: list[str] = []
        if self_phone:
            phone_lines.append(f"Self: {self_phone}")
        if parent_phone:
            phone_lines.append(f"Parent: {parent_phone}")
        if self_email:
            email_lines.append(f"Self: {self_email}")
        if parent_email:
            email_lines.append(f"Parent: {parent_email}")
        consent_lines = [
            f"Self: {consent_text(has_self_consent)}",
            f"Parent: {consent_text(has_parent_consent)}",
        ]
        return ContactDisplay(
            phone=join_lines(phone_lines),
            email=join_lines(email_lines),
            consent=join_lines(consent_lines),
            no_consent_reason=" and ".join(no_consent_parts) + " contact details do not permit onward sharing"
            if no_consent_parts
            else "",
            cross_family=cross_family,
            cross_family_reason=cross_family_reason,
        )

    if self_visible:
        return ContactDisplay(
            phone=self_phone,
            email=self_email,
            consent=consent_text(has_self_consent),
            no_consent_reason="self contact details do not permit onward sharing"
            if not has_self_consent
            else "",
            cross_family=cross_family,
            cross_family_reason=cross_family_reason,
        )

    if parent_visible:
        return ContactDisplay(
            phone=parent_phone,
            email=parent_email,
            consent=consent_text(has_parent_consent),
            no_consent_reason="parent contact details do not permit onward sharing"
            if not has_parent_consent
            else "",
            cross_family=cross_family,
            cross_family_reason=cross_family_reason,
        )

    fallback_has_context = bool(
        self_contact is not None
        or main_contact is not None
        or member_candidates
        or signup_candidates
    )
    return ContactDisplay(
        phone="",
        email="",
        consent="No" if fallback_has_context else "",
        no_consent_reason="",
        cross_family=cross_family,
        cross_family_reason=cross_family_reason,
    )


def email_matches_name(candidate: ContactCandidate, source_name: str) -> bool:
    first_name, last_name = split_person_name(source_name)
    norm_first = normalize_name(first_name)
    norm_last = normalize_name(last_name)
    local = normalize_email_local(candidate.email)
    if not norm_first or not norm_last or not local:
        return False
    return norm_first in local and norm_last in local


def contact_quality_score(candidate: ContactCandidate) -> tuple[int, int, int]:
    return (
        1 if candidate.email else 0,
        1 if candidate.mobile else 0,
        1 if candidate.phone else 0,
    )


def unique_best_contact(candidates: list[ContactCandidate]) -> list[ContactCandidate]:
    if len(candidates) <= 1:
        return candidates
    ranked = sorted(candidates, key=contact_quality_score, reverse=True)
    best = contact_quality_score(ranked[0])
    second = contact_quality_score(ranked[1])
    if best > second:
        return [ranked[0]]
    return candidates


def disambiguate_contact_candidates(
    contact_candidates: list[ContactCandidate],
    source_name: str,
) -> tuple[list[ContactCandidate], bool]:
    if len(contact_candidates) <= 1:
        return contact_candidates, False

    active = [candidate for candidate in contact_candidates if candidate.member_status == "Active Member"]
    if len(active) == 1:
        return active, True
    best_active = unique_best_contact(active)
    if len(best_active) == 1:
        return best_active, True

    matching_active = [candidate for candidate in active if email_matches_name(candidate, source_name)]
    if len(matching_active) == 1:
        return matching_active, True

    non_lapsed = [
        candidate
        for candidate in contact_candidates
        if candidate.member_status not in {"Lapsed Member", "Non Member", ""}
    ]
    if len(non_lapsed) == 1:
        return non_lapsed, True
    best_non_lapsed = unique_best_contact(non_lapsed)
    if len(best_non_lapsed) == 1:
        return best_non_lapsed, True

    matching_non_lapsed = [
        candidate for candidate in non_lapsed if email_matches_name(candidate, source_name)
    ]
    if len(matching_non_lapsed) == 1:
        return matching_non_lapsed, True

    matching_all = [candidate for candidate in contact_candidates if email_matches_name(candidate, source_name)]
    if len(matching_all) == 1:
        return matching_all, True

    return contact_candidates, False


def nickname_variants(first_name: str) -> set[str]:
    norm_first = normalize_name(first_name)
    variants = {norm_first}
    if norm_first in NICKNAME_OVERRIDES:
        variants |= NICKNAME_OVERRIDES[norm_first]
    return variants


def unique_nickname_candidates(name: str, by_name: dict[str, list[object]]) -> list[object]:
    first_name, last_name = split_person_name(name)
    norm_last = normalize_name(last_name)
    if not norm_last:
        return []
    keys = []
    for variant in nickname_variants(first_name):
        candidate_key = normalize_name(f"{variant} {norm_last}")
        if candidate_key in by_name:
            keys.append(candidate_key)
    keys = sorted(set(keys))
    if len(keys) != 1:
        return []
    candidates = by_name[keys[0]]
    return candidates if len(candidates) == 1 else []


def fuzzy_score_parts(source: str, target: str) -> float:
    return SequenceMatcher(None, normalize_name(source), normalize_name(target)).ratio()


def unique_fuzzy_candidates(name: str, by_name: dict[str, list[object]]) -> list[object]:
    first_name, last_name = split_person_name(name)
    norm_first = normalize_name(first_name)
    norm_last = normalize_name(last_name)
    if not norm_first or not norm_last:
        return []

    matches: list[tuple[float, str, list[object]]] = []
    for key, candidates in by_name.items():
        if len(candidates) != 1:
            continue
        candidate = candidates[0]
        cand_first = normalize_name(candidate.first_name)
        cand_last = normalize_name(candidate.last_name)
        if not cand_first or not cand_last:
            continue
        if cand_first[:1] != norm_first[:1] or cand_last[:1] != norm_last[:1]:
            continue
        first_score = fuzzy_score_parts(first_name, candidate.first_name)
        last_score = fuzzy_score_parts(last_name, candidate.last_name)
        full_score = fuzzy_score_parts(
            f"{first_name} {last_name}",
            f"{candidate.first_name} {candidate.last_name}",
        )
        if first_score >= 0.72 and last_score >= 0.80 and full_score >= 0.84:
            matches.append((full_score + last_score + first_score, key, candidates))

    if not matches:
        return []

    matches.sort(key=lambda item: item[0], reverse=True)
    best_score, best_key, best_candidates = matches[0]
    if len(matches) > 1 and abs(best_score - matches[1][0]) < 0.08:
        return []
    return best_candidates


def resolve_row(
    name: str,
    members_by_name: dict[str, list[MemberCandidate]],
    contacts_by_name: dict[str, list[ContactCandidate]],
    signups_by_name: dict[str, list[SignupCandidate]],
    junior_main_contacts_by_name: dict[str, list[JuniorMainContactCandidate]],
) -> dict[str, str]:
    norm_name = normalize_name(name)
    override_name = next(iter(EXACT_NAME_OVERRIDES.get(norm_name, set())), "")
    override_applied = bool(override_name)
    if override_name:
        norm_name = override_name
    member_candidates = members_by_name.get(norm_name, [])
    contact_candidates, best_fit_applied = disambiguate_contact_candidates(contacts_by_name.get(norm_name, []), name)
    signup_candidates = signups_by_name.get(norm_name, [])
    junior_main_contact_candidates = junior_main_contacts_by_name.get(norm_name, [])

    contact_display = choose_contact_detail(
        member_candidates,
        contact_candidates,
        signup_candidates,
        junior_main_contact_candidates,
        all_contact_candidates=contacts_by_name.get(norm_name, []),
    )

    if len(member_candidates) == 1:
        result = {
            "category": member_candidates[0].category,
            "phone": contact_display.phone,
            "email": contact_display.email,
            "consent": contact_display.consent,
            "match_note": "Override" if override_applied else ("Family" if contact_display.cross_family else ""),
            "review_section": "explicit_override" if override_applied else ("cross_family" if contact_display.cross_family else ""),
            "review_target": candidate_full_name(member_candidates[0]) if override_applied else "",
            "review_reason": "explicit full-name override from name_overrides.csv" if override_applied else (contact_display.cross_family_reason if contact_display.cross_family else ""),
            "no_consent_reason": contact_display.no_consent_reason,
        }
        if not contact_display.phone and not contact_display.email:
            result["no_contact_info"] = True
        return result

    if len(member_candidates) > 1:
        return {"category": "No Match", "phone": "", "email": "", "consent": "", "match_note": "", "no_consent_reason": ""}

    if len(signup_candidates) == 1:
        return {
            "category": signup_candidates[0].category,
            "phone": contact_display.phone,
            "email": contact_display.email,
            "consent": contact_display.consent,
            "match_note": "Override" if override_applied else "",
            "review_section": "explicit_override" if override_applied else "",
            "review_target": candidate_full_name(signup_candidates[0]) if override_applied else "",
            "review_reason": "explicit full-name override from name_overrides.csv" if override_applied else "",
            "no_consent_reason": contact_display.no_consent_reason,
        }

    if len(signup_candidates) > 1:
        return {"category": "No Match", "phone": "", "email": "", "consent": "", "match_note": "", "no_consent_reason": ""}

    if len(contact_candidates) == 1:
        return {
            "category": "Not Signed Up",
            "phone": contact_display.phone,
            "email": contact_display.email,
            "consent": contact_display.consent,
            "match_note": "Override" if override_applied else ("Best Fit" if best_fit_applied else ""),
            "review_section": "explicit_override" if override_applied else ("not_signed_up" if not override_applied else ""),
            "review_target": candidate_full_name(contact_candidates[0]) if override_applied else "",
            "review_reason": "explicit full-name override from name_overrides.csv" if override_applied else ("contact-only match - no current membership" if not override_applied else ""),
            "no_consent_reason": contact_display.no_consent_reason,
            "review_not_signed_up": True,
        }

    member_candidates = unique_nickname_candidates(name, members_by_name)
    contact_candidates, best_fit_applied = disambiguate_contact_candidates(
        unique_nickname_candidates(name, contacts_by_name),
        name,
    )
    signup_candidates = unique_nickname_candidates(name, signups_by_name)
    junior_main_contact_candidates = unique_nickname_candidates(name, junior_main_contacts_by_name)
    contact_display = choose_contact_detail(
        member_candidates,
        contact_candidates,
        signup_candidates,
        junior_main_contact_candidates,
    )

    if len(member_candidates) == 1:
        return {
            "category": member_candidates[0].category,
            "phone": contact_display.phone,
            "email": contact_display.email,
            "consent": contact_display.consent,
            "match_note": "Nickname",
            "review_section": "nickname",
            "review_target": candidate_full_name(member_candidates[0]),
            "review_reason": "nickname mapping from nicknames.csv",
            "no_consent_reason": contact_display.no_consent_reason,
        }

    if len(signup_candidates) == 1:
        return {
            "category": signup_candidates[0].category,
            "phone": contact_display.phone,
            "email": contact_display.email,
            "consent": contact_display.consent,
            "match_note": "Nickname",
            "review_section": "nickname",
            "review_target": candidate_full_name(signup_candidates[0]),
            "review_reason": "nickname mapping from nicknames.csv",
            "no_consent_reason": contact_display.no_consent_reason,
        }

    if len(contact_candidates) == 1:
        return {
            "category": "Not Signed Up",
            "phone": contact_display.phone,
            "email": contact_display.email,
            "consent": contact_display.consent,
            "match_note": "Best Fit" if best_fit_applied else "Nickname",
            "review_section": "nickname",
            "review_target": candidate_full_name(contact_candidates[0]),
            "review_reason": "nickname mapping from nicknames.csv",
            "no_consent_reason": contact_display.no_consent_reason,
        }

    member_candidates = unique_fuzzy_candidates(name, members_by_name)
    contact_candidates, best_fit_applied = disambiguate_contact_candidates(
        unique_fuzzy_candidates(name, contacts_by_name),
        name,
    )
    signup_candidates = unique_fuzzy_candidates(name, signups_by_name)
    junior_main_contact_candidates = unique_fuzzy_candidates(name, junior_main_contacts_by_name)
    contact_display = choose_contact_detail(
        member_candidates,
        contact_candidates,
        signup_candidates,
        junior_main_contact_candidates,
    )

    if len(member_candidates) == 1:
        return {
            "category": member_candidates[0].category,
            "phone": contact_display.phone,
            "email": contact_display.email,
            "consent": contact_display.consent,
            "match_note": "Fuzzy",
            "review_section": "fuzzy",
            "review_target": candidate_full_name(member_candidates[0]),
            "review_reason": "fuzzy name match",
            "no_consent_reason": contact_display.no_consent_reason,
        }

    if len(signup_candidates) == 1:
        return {
            "category": signup_candidates[0].category,
            "phone": contact_display.phone,
            "email": contact_display.email,
            "consent": contact_display.consent,
            "match_note": "Fuzzy",
            "review_section": "fuzzy",
            "review_target": candidate_full_name(signup_candidates[0]),
            "review_reason": "fuzzy name match",
            "no_consent_reason": contact_display.no_consent_reason,
        }

    if len(contact_candidates) == 1:
        return {
            "category": "Not Signed Up",
            "phone": contact_display.phone,
            "email": contact_display.email,
            "consent": contact_display.consent,
            "match_note": "Best Fit" if best_fit_applied else "Fuzzy",
            "review_section": "fuzzy",
            "review_target": candidate_full_name(contact_candidates[0]),
            "review_reason": "fuzzy name match",
            "no_consent_reason": contact_display.no_consent_reason,
        }

    return {
        "category": "No Match",
        "phone": "",
        "email": "",
        "consent": "",
        "match_note": "",
        "review_section": "no_match",
        "review_target": "",
        "review_reason": "no unique candidate",
        "no_consent_reason": "",
    }


def build_team_rows(
    teams: dict[str, list[dict[str, object]]],
    members_by_name: dict[str, list[MemberCandidate]],
    contacts_by_name: dict[str, list[ContactCandidate]],
    signups_by_name: dict[str, list[SignupCandidate]],
    junior_main_contacts_by_name: dict[str, list[JuniorMainContactCandidate]],
) -> tuple[dict[str, list[dict[str, str]]], list[ReviewEntry]]:
    resolved: dict[str, list[dict[str, str]]] = {}
    review_entries: list[ReviewEntry] = []

    for team_name, entries in teams.items():
        rows = []
        for entry in entries:
            details = resolve_row(
                entry["name"],
                members_by_name,
                contacts_by_name,
                signups_by_name,
                junior_main_contacts_by_name,
            )
            rows.append(
                {
                    "captain": "C" if entry["captain"] else "",
                    "name": str(entry["name"]),
                    "match_note": details["match_note"],
                    "category": details["category"],
                    "consent": details["consent"],
                    "phone": details["phone"],
                    "email": details["email"],
                }
            )
            if details.get("review_section"):
                review_entries.append(
                    ReviewEntry(
                        section=details["review_section"],
                        source_name=str(entry["name"]),
                        target_name=details.get("review_target", ""),
                        reason=details.get("review_reason", ""),
                        team_name=team_name.title(),
                    )
                )
            if details.get("no_consent_reason"):
                review_entries.append(
                    ReviewEntry(
                        section="no_consent",
                        source_name=str(entry["name"]),
                        target_name="",
                        reason=details["no_consent_reason"],
                        team_name=team_name.title(),
                    )
                )
            if details.get("no_contact_info"):
                review_entries.append(
                    ReviewEntry(
                        section="no_contact_info",
                        source_name=str(entry["name"]),
                        target_name="",
                        reason="matched but no email or phone in club database",
                        team_name=team_name.title(),
                    )
                )

        rows.sort(key=lambda row: (0 if row["captain"] else 1, first_name_key(row["name"]), row["name"].casefold()))
        resolved[team_name] = rows
    return resolved, review_entries


def group_review_entries(review_entries: list[ReviewEntry]) -> dict[str, list[dict[str, object]]]:
    grouped: dict[tuple[str, str], dict[str, object]] = {}
    for entry in review_entries:
        key = (entry.section, normalize_name(entry.source_name))
        if key not in grouped:
            grouped[key] = {
                "display_name": entry.source_name,
                "target_name": entry.target_name,
                "reason": entry.reason,
                "teams": set(),
            }
        grouped[key]["display_name"] = choose_display_name(grouped[key]["display_name"], entry.source_name)
        grouped[key]["teams"].add(entry.team_name)
        if entry.target_name:
            grouped[key]["target_name"] = entry.target_name

    ordered_sections = ["no_match", "no_contact_info", "not_signed_up", "no_consent", "explicit_override", "cross_family", "nickname", "fuzzy"]
    result: dict[str, list[dict[str, object]]] = {section: [] for section in ordered_sections}
    for (section, _), value in grouped.items():
        result[section].append(value)
    for section in ordered_sections:
        result[section].sort(key=lambda item: normalize_name(item["display_name"]))
    return result


def review_markdown_text(review_groups: dict[str, list[dict[str, object]]]) -> str:
    lines = [
        "# Team Match Review",
        "",
        "This file records the current non-trivial name matching used by [generate_team_contact_lists.py](./generate_team_contact_lists.py).",
        "",
        "## Remaining No Match Names",
        "",
    ]
    for item in review_groups["no_match"]:
        teams = ", ".join(sorted(item["teams"]))
        lines.append(f"- `{item['display_name']}` - {item['reason']}. Teams: {teams}")

    lines.extend(["", "## No Contact Information (Matched but No Contact Details)", ""])
    for item in review_groups.get("no_contact_info", []):
        teams = ", ".join(sorted(item["teams"]))
        lines.append(f"- `{item['display_name']}` - {item['reason']}. Teams: {teams}")

    lines.extend(["", "## Not Signed Up (Contact Only, No Current Membership)", ""])
    for item in review_groups.get("not_signed_up", []):
        teams = ", ".join(sorted(item["teams"]))
        lines.append(
            f"- `{item['display_name']}` - {item['reason']}. Teams: {teams}"
        )

    lines.extend(["", "## No Consent Contact Details", ""])
    for item in review_groups["no_consent"]:
        teams = ", ".join(sorted(item["teams"]))
        lines.append(f"- `{item['display_name']}` - {item['reason']}. Teams: {teams}")

    lines.extend(["", "## Explicit Full-Name Overrides", ""])
    for item in review_groups["explicit_override"]:
        teams = ", ".join(sorted(item["teams"]))
        lines.append(
            f"- `{item['display_name']}` -> `{item['target_name']}` - {item['reason']}. Teams: {teams}"
        )

    lines.extend(["", "## Cross-Family Contact Matches", ""])
    for item in review_groups.get("cross_family", []):
        teams = ", ".join(sorted(item["teams"]))
        lines.append(
            f"- `{item['display_name']}` -> `{item['target_name']}` - {item['reason']}. Teams: {teams}"
        )

    lines.extend(["", "## Not Signed Up (Contact Only, No Current Membership)", ""])
    for item in review_groups.get("not_signed_up", []):
        teams = ", ".join(sorted(item["teams"]))
        lines.append(
            f"- `{item['display_name']}` - {item['reason']}. Teams: {teams}"
        )

    lines.extend(["", "## Short-Name / Nickname Matches", ""])
    for item in review_groups["nickname"]:
        teams = ", ".join(sorted(item["teams"]))
        lines.append(
            f"- `{item['display_name']}` -> `{item['target_name']}` - {item['reason']}. Teams: {teams}"
        )

    lines.extend(["", "## Fuzzy Matches", ""])
    for item in review_groups["fuzzy"]:
        teams = ", ".join(sorted(item["teams"]))
        lines.append(
            f"- `{item['display_name']}` -> `{item['target_name']}` - {item['reason']}. Teams: {teams}"
        )

    lines.append("")
    return "\n".join(lines)


def write_review_markdown(review_groups: dict[str, list[dict[str, object]]], output_path: Path) -> None:
    output_path.write_text(review_markdown_text(review_groups), encoding="utf-8")


def write_nickname_csv(review_groups: dict[str, list[dict[str, object]]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["source_name", "target_name", "reason", "teams"])
        for item in review_groups["nickname"]:
            writer.writerow([
                item["display_name"],
                item["target_name"],
                item["reason"],
                ", ".join(sorted(item["teams"])),
            ])


def write_review_pdf(review_groups: dict[str, list[dict[str, object]]], output_path: Path) -> None:
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReviewTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=16,
        textColor=colors.HexColor(f"#{THEME['navy']}"),
        spaceAfter=8,
    )
    section_style = ParagraphStyle(
        "ReviewSection",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=12,
        textColor=colors.HexColor(f"#{THEME['gold']}"),
        spaceBefore=8,
        spaceAfter=4,
    )
    body_style = ParagraphStyle(
        "ReviewBody",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10,
        textColor=colors.HexColor(f"#{THEME['ink']}"),
        leading=13,
        spaceAfter=3,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    story = [
        Paragraph("Avondale Tennis Club | Team Match Review", title_style),
        Paragraph(f"Generated {datetime.now().strftime('%d %b %Y %H:%M')}", body_style),
    ]

    section_map = [
        ("Remaining No Match Names", "no_match"),
        ("No Contact Information (Matched but No Contact Details)", "no_contact_info"),
        ("Not Signed Up (Contact Only, No Current Membership)", "not_signed_up"),
        ("No Consent Contact Details", "no_consent"),
        ("Explicit Full-Name Overrides", "explicit_override"),
        ("Cross-Family Contact Matches", "cross_family"),
        ("Short-Name / Nickname Matches", "nickname"),
        ("Fuzzy Matches", "fuzzy"),
    ]
    for title, key in section_map:
        story.append(Spacer(1, 2 * mm))
        story.append(Paragraph(title, section_style))
        for item in review_groups.get(key, []):
            teams = ", ".join(sorted(item["teams"]))
            if item["target_name"]:
                text = f"<b>{item['display_name']}</b> -> <b>{item['target_name']}</b> - {item['reason']}. Teams: {teams}"
            else:
                text = f"<b>{item['display_name']}</b> - {item['reason']}. Teams: {teams}"
            story.append(Paragraph(text, body_style))

    doc.build(story)


def write_workbook(title: str, teams: dict[str, list[dict[str, str]]], output_path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    title_fill = PatternFill("solid", fgColor=THEME["navy"])
    header_fill = PatternFill("solid", fgColor=THEME["light"])
    gold_fill = PatternFill("solid", fgColor=THEME["gold"])
    white_font = Font(color=THEME["white"], bold=True, size=14)
    header_font = Font(color=THEME["ink"], bold=True)
    bold_font = Font(color=THEME["ink"], bold=True)
    body_font = Font(color=THEME["ink"])
    footnote_font = Font(color=THEME["ink"], italic=True, size=10)
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for team_name, rows in teams.items():
        ws = workbook.create_sheet(sheet_title(team_name))
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A4"
        ws.merge_cells("A1:G1")
        ws["A1"] = f"Avondale Tennis Club | {title} | {team_name.title()}"
        ws["A1"].fill = title_fill
        ws["A1"].font = white_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24

        ws.merge_cells("A2:G2")
        ws["A2"] = f"Generated {datetime.now().strftime('%d %b %Y %H:%M')}"
        ws["A2"].fill = gold_fill
        ws["A2"].font = Font(color=THEME["ink"], bold=True)
        ws["A2"].alignment = Alignment(horizontal="center")

        headers = ["Captain", "Name", "Category", "Consent", "Phone", "Email", "Match"]
        for col_index, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_index, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        row_index = 4
        for item in rows:
            values = [item["captain"], item["name"], item["category"], item["consent"], item["phone"], item["email"], item["match_note"]]
            for col_index, value in enumerate(values, start=1):
                cell = ws.cell(row=row_index, column=col_index, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.font = bold_font if item["captain"] == "C" else body_font
            row_index += 1

        row_index += 1
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=7)
        ws.cell(row=row_index, column=1, value="* Consent: No means the captain may use the phone/email for team management, but may not pass it on to anyone else.").font = footnote_font
        row_index += 1
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=7)
        ws.cell(row=row_index, column=1, value="** Not Signed Up = no current membership was found.").font = footnote_font
        row_index += 1
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=7)
        ws.cell(row=row_index, column=1, value="*** No Match = no unique exact-name match was found in current club member/contact records.").font = footnote_font
        row_index += 1
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=7)
        ws.cell(row=row_index, column=1, value="**** Best Fit = one row was chosen from several exact-name candidates using the best available contact evidence.").font = footnote_font
        row_index += 1
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=7)
        ws.cell(row=row_index, column=1, value="***** Override = resolved by an explicit override; this usually represents a typo or misspelling in the source data.").font = footnote_font
        row_index += 1
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=7)
        ws.cell(row=row_index, column=1, value="****** Nickname = resolved by a short-name or nickname rule; this may also represent a typo or misspelling in the source data.").font = footnote_font
        row_index += 1
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=7)
        ws.cell(row=row_index, column=1, value="******* Fuzzy = resolved by the cautious fuzzy fallback; this represents a typo or misspelling in the source data.").font = footnote_font

        widths = {"A": 10, "B": 28, "C": 18, "D": 15, "E": 18, "F": 34, "G": 12}
        for column, width in widths.items():
            ws.column_dimensions[column].width = width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def write_csvs(teams: dict[str, list[dict[str, str]]], output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    for team_name, rows in teams.items():
        slug = re.sub(r"[^a-z0-9]+", "-", team_name.casefold()).strip("-")
        csv_path = output_dir / f"{slug}.csv"
        with csv_path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(["Captain", "Name", "Category", "Consent", "Phone", "Email", "Match"])
            for row in rows:
                writer.writerow([row["captain"], row["name"], row["category"], row["consent"], row["phone"], row["email"], row["match_note"]])
            writer.writerow([])
            writer.writerow(["* Consent: No means the captain may use the phone/email for team management, but may not pass it on to anyone else."])
            writer.writerow(["** Not Signed Up = no current membership was found."])
            writer.writerow(["*** No Match = no unique exact-name match was found in current club member/contact records."])
            writer.writerow(["**** Best Fit = one row was chosen from several exact-name candidates using the best available contact evidence."])
            writer.writerow(["***** Override = resolved by an explicit override; this usually represents a typo or misspelling in the source data."])
            writer.writerow(["****** Nickname = resolved by a short-name or nickname rule; this may also represent a typo or misspelling in the source data."])
            writer.writerow(["******* Fuzzy = resolved by the cautious fuzzy fallback; this represents a typo or misspelling in the source data."])


def write_captain_email_list(all_teams: dict[str, dict[str, list[dict[str, str]]]], output_path: Path) -> None:
    emails: list[str] = []
    seen: set[str] = set()

    for teams in all_teams.values():
        for rows in teams.values():
            for row in rows:
                if row["captain"] != "C":
                    continue
                raw_email = row["email"].strip()
                if not raw_email:
                    continue
                for part in raw_email.splitlines():
                    email = re.sub(r"^(Self|Parent):\s*", "", part).strip()
                    if not email or email.casefold() == "no consent":
                        continue
                    key = email.casefold()
                    if key in seen:
                        continue
                    seen.add(key)
                    emails.append(email)

    output_path.write_text(", ".join(emails) + ("\n" if emails else ""), encoding="utf-8")


def extract_preferred_email(value: str) -> str:
    lines = [line.strip() for line in str(value or "").splitlines() if line.strip()]
    cleaned = []
    for line in lines:
        line = re.sub(r"^(Self|Parent):\s*", "", line, flags=re.I).strip()
        if line and "@" in line:
            cleaned.append(line)
    if not cleaned:
        return ""
    self_email = next((line for line in cleaned if normalize_email_local(line)), "")
    return self_email or cleaned[0]


def section_label_from_title(title: str) -> str:
    first_word = title.split()[0].strip()
    return first_word.title()


def attachment_file_name(source_doc: str, team_name: str) -> str:
    return f"{source_doc} - {team_slug(team_name)}.pdf"


def attachment_kind_label(attachment: CaptainEmailAttachment) -> str:
    if attachment.kind == "own":
        return f"your team: {attachment.team_name}"
    if attachment.kind == "next":
        return f"the next team down: {attachment.team_name}"
    if attachment.kind == "reserves":
        return "reserves"
    return attachment.team_name


def attachment_mode_summary(mode: str) -> str:
    summaries = {
        "own-plus-reserves": "Each captain receives their own team sheet plus the reserves sheet.",
        "own-next-plus-reserves": "Each captain receives their own team sheet, the next team down, and the reserves sheet.",
        "all-in-section": "Each captain receives every team sheet in their section, including reserves where present.",
    }
    return summaries[mode]


def build_captain_email_attachments(
    source_doc: str,
    team_name: str,
    next_team_name: str,
    reserves_name: str,
    all_team_names: list[str],
    attachment_mode: str,
) -> list[CaptainEmailAttachment]:
    attachment_plan: list[tuple[str, str]] = []
    if attachment_mode == "own-plus-reserves":
        attachment_plan.append(("own", team_name))
        if reserves_name:
            attachment_plan.append(("reserves", reserves_name))
    elif attachment_mode == "own-next-plus-reserves":
        attachment_plan.append(("own", team_name))
        if next_team_name:
            attachment_plan.append(("next", next_team_name))
        if reserves_name:
            attachment_plan.append(("reserves", reserves_name))
    elif attachment_mode == "all-in-section":
        attachment_plan.extend(("section", candidate) for candidate in all_team_names if candidate)
    else:
        raise ValueError(f"Unsupported attachment mode: {attachment_mode}")

    attachments: list[CaptainEmailAttachment] = []
    seen_files: set[str] = set()
    for kind, candidate_name in attachment_plan:
        if not candidate_name:
            continue
        file_name = attachment_file_name(source_doc, candidate_name)
        if file_name in seen_files:
            continue
        seen_files.add(file_name)
        attachments.append(
            CaptainEmailAttachment(
                team_name=candidate_name.title(),
                file_name=file_name,
                kind=kind,
            )
        )
    return attachments


def build_captain_email_jobs(
    source_doc: str,
    title: str,
    teams: dict[str, list[dict[str, str]]],
    attachment_mode: str,
) -> list[CaptainEmailJob]:
    jobs: list[CaptainEmailJob] = []
    team_names = list(teams.keys())
    reserves_name = next((name for name in team_names if name.casefold() == "reserves"), "")
    section = section_label_from_title(title)

    ordered_non_reserves = [name for name in team_names if name.casefold() != "reserves"]
    for index, team_name in enumerate(ordered_non_reserves):
        rows = teams[team_name]
        captain_row = next((row for row in rows if row["captain"] == "C"), None)
        if not captain_row:
            continue
        next_team_name = ordered_non_reserves[index + 1] if index + 1 < len(ordered_non_reserves) else reserves_name
        attachments = build_captain_email_attachments(
            source_doc=source_doc,
            team_name=team_name,
            next_team_name=next_team_name,
            reserves_name=reserves_name,
            all_team_names=team_names,
            attachment_mode=attachment_mode,
        )
        captain_email = extract_preferred_email(captain_row["email"])
        blocked_reasons = []
        if not captain_email:
            blocked_reasons.append("No captain email address available")
        if not attachments:
            blocked_reasons.append("No attachments available for this team")
        can_send = not blocked_reasons
        blocked_reason = "; ".join(blocked_reasons)
        jobs.append(
            CaptainEmailJob(
                source_doc=source_doc,
                section=section,
                team_name=team_name.title(),
                next_team_name=next_team_name.title() if next_team_name else "",
                captain_name=captain_row["name"],
                captain_email=captain_email,
                attachment_mode=attachment_mode,
                attachments=attachments,
                can_send=can_send,
                blocked_reason=blocked_reason,
            )
        )
    return jobs


def write_captain_email_jobs_json(
    jobs: list[CaptainEmailJob], output_path: Path, attachment_mode: str
) -> None:
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "attachment_mode": attachment_mode,
        "jobs": [
            {
                "source_doc": job.source_doc,
                "section": job.section,
                "team_name": job.team_name,
                "next_team_name": job.next_team_name,
                "captain_name": job.captain_name,
                "captain_email": job.captain_email,
                "attachment_mode": job.attachment_mode,
                "attachments": [
                    {
                        "team_name": attachment.team_name,
                        "file_name": attachment.file_name,
                        "kind": attachment.kind,
                    }
                    for attachment in job.attachments
                ],
                "can_send": job.can_send,
                "blocked_reason": job.blocked_reason,
            }
            for job in jobs
        ],
    }
    output_path.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")


def write_captain_email_jobs_csv(jobs: list[CaptainEmailJob], output_path: Path) -> None:
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow([
            "section",
            "team_name",
            "next_team_name",
            "captain_name",
            "captain_email",
            "attachment_mode",
            "attachment_count",
            "attachments",
            "can_send",
            "blocked_reason",
        ])
        for job in jobs:
            writer.writerow([
                job.section,
                job.team_name,
                job.next_team_name,
                job.captain_name,
                job.captain_email,
                job.attachment_mode,
                len(job.attachments),
                "; ".join(attachment.file_name for attachment in job.attachments),
                "yes" if job.can_send else "no",
                job.blocked_reason,
            ])


def write_captain_email_send_list(
    jobs: list[CaptainEmailJob], output_path: Path, attachment_mode: str
) -> None:
    lines = [
        "# Team Captain Email Send List",
        "",
        f"Attachment mode: `{attachment_mode}`",
        CAPTAIN_ATTACHMENT_MODE_LABELS[attachment_mode],
        "",
        attachment_mode_summary(attachment_mode),
        "",
    ]
    by_section: dict[str, list[CaptainEmailJob]] = defaultdict(list)
    for job in jobs:
        by_section[job.section].append(job)
    for section in sorted(by_section.keys()):
        lines.extend([f"## {section}", ""])
        for job in by_section[section]:
            email_text = f"`{job.captain_email}`" if job.captain_email else f"`{job.blocked_reason or 'No email'}`"
            lines.append(f"- {job.captain_name} — {email_text}")
            lines.append("  Attach:")
            for attachment in job.attachments:
                attachment_link = (
                    f"[{attachment.file_name}]"
                    f"(/mnt/c/dev/avondale-n8n/Teams/generated/{attachment.file_name.replace(' ', '%20')})"
                )
                lines.append(
                    "  - "
                    f"{attachment_kind_label(attachment)} — "
                    f"{attachment_link}"
                )
            if not job.can_send:
                lines.append(f"  Blocked: {job.blocked_reason}")
        lines.append("")
    output_path.write_text("\n".join(lines).strip() + "\n", encoding="utf-8")


def team_slug(team_name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", team_name.casefold()).strip("-")

def write_pdf(title: str, teams: dict[str, list[dict[str, str]]], output_path: Path) -> None:
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "AvondaleTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=16,
        textColor=colors.HexColor(f"#{THEME['navy']}"),
        spaceAfter=8,
    )
    subtitle_style = ParagraphStyle(
        "AvondaleSubtitle",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10,
        textColor=colors.HexColor(f"#{THEME['gold']}"),
        spaceAfter=10,
    )
    note_style = ParagraphStyle(
        "AvondaleNote",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.HexColor(f"#{THEME['ink']}"),
        leading=11,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )

    story = []
    team_names = list(teams.keys())
    for index, team_name in enumerate(team_names):
        rows = teams[team_name]
        story.append(Paragraph(f"Avondale Tennis Club | {title} | {team_name.title()}", title_style))
        story.append(Paragraph(f"Generated {datetime.now().strftime('%d %b %Y %H:%M')}", subtitle_style))

        table_data = [["Captain", "Name", "Category", "Consent", "Phone", "Email", "Match"]]
        for row in rows:
            table_data.append([row["captain"], row["name"], row["category"], row["consent"], row["phone"], row["email"], row["match_note"]])

        table = Table(table_data, colWidths=[18 * mm, 48 * mm, 28 * mm, 20 * mm, 30 * mm, 60 * mm, 20 * mm], repeatRows=1)
        style_commands = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{THEME['navy']}")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D0D7DE")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#FFFFFF"), colors.HexColor(f"#{THEME['light']}")]),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ]
        for row_number, row in enumerate(rows, start=1):
            if row["captain"] == "C":
                style_commands.append(("FONTNAME", (0, row_number), (-1, row_number), "Helvetica-Bold"))
        table.setStyle(TableStyle(style_commands))
        story.append(table)
        story.append(Spacer(1, 6 * mm))
        story.append(Paragraph("* Consent: No means the captain may use the phone/email for team management, but may not pass it on to anyone else.", note_style))
        story.append(Paragraph("** Not Signed Up = no current membership was found.", note_style))
        story.append(Paragraph("*** No Match = no unique exact-name match was found in current club member/contact records.", note_style))
        story.append(Paragraph("**** Best Fit = one row was chosen from several exact-name candidates using the best available contact evidence.", note_style))
        story.append(Paragraph("***** Override = resolved by an explicit override; this usually represents a typo or misspelling in the source data.", note_style))
        story.append(Paragraph("****** Nickname = resolved by a short-name or nickname rule; this may also represent a typo or misspelling in the source data.", note_style))
        story.append(Paragraph("******* Fuzzy = resolved by the cautious fuzzy fallback; this represents a typo or misspelling in the source data.", note_style))

        if index < len(team_names) - 1:
            story.append(PageBreak())

    doc.build(story)


def main() -> None:
    args = parse_args()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    members_by_name, contacts_by_name, signups_by_name, junior_main_contacts_by_name = load_lookup_data()
    all_review_entries: list[ReviewEntry] = []
    all_resolved_teams: dict[str, dict[str, list[dict[str, str]]]] = {}
    all_captain_jobs: list[CaptainEmailJob] = []

    for docx_path in sorted(BASE_DIR.glob("*.docx")):
        title = doc_title_from_filename(docx_path)
        teams = parse_docx_teams(docx_path)
        resolved_teams, review_entries = build_team_rows(
            teams,
            members_by_name,
            contacts_by_name,
            signups_by_name,
            junior_main_contacts_by_name,
        )
        all_review_entries.extend(review_entries)
        all_resolved_teams[docx_path.stem] = resolved_teams
        all_captain_jobs.extend(
            build_captain_email_jobs(
                docx_path.stem,
                title,
                resolved_teams,
                args.captain_attachment_mode,
            )
        )

        workbook_path = OUTPUT_DIR / f"{docx_path.stem} - Contact Lists.xlsx"
        csv_dir = OUTPUT_DIR / f"{docx_path.stem} - CSV"
        old_pdf_path = OUTPUT_DIR / f"{docx_path.stem} - Contact Lists.pdf"

        write_workbook(title, resolved_teams, workbook_path)
        write_csvs(resolved_teams, csv_dir)
        for team_name, rows in resolved_teams.items():
            write_pdf(
                title,
                {team_name: rows},
                OUTPUT_DIR / f"{docx_path.stem} - {team_slug(team_name)}.pdf",
            )
        old_pdf_path.unlink(missing_ok=True)

        no_match_count = sum(1 for rows in resolved_teams.values() for row in rows if row["category"] == "No Match")
        not_signed_count = sum(1 for rows in resolved_teams.values() for row in rows if row["category"] == "Not Signed Up")
        print(f"{docx_path.name}: sheets={len(resolved_teams)} no_match={no_match_count} not_signed_up={not_signed_count}")

    review_groups = group_review_entries(all_review_entries)
    write_captain_email_list(all_resolved_teams, OUTPUT_DIR / "team-captains-email-list.txt")
    write_captain_email_jobs_json(
        all_captain_jobs,
        OUTPUT_DIR / "team-captain-email-jobs.json",
        args.captain_attachment_mode,
    )
    write_captain_email_jobs_csv(all_captain_jobs, OUTPUT_DIR / "team-captain-email-jobs.csv")
    write_captain_email_send_list(
        all_captain_jobs,
        OUTPUT_DIR / "CAPTAIN_EMAIL_SEND_LIST.md",
        args.captain_attachment_mode,
    )
    write_nickname_csv(review_groups, OUTPUT_DIR / "nickname-matches.csv")
    write_review_markdown(review_groups, BASE_DIR / "NO_MATCH_NAMES.md")
    write_review_markdown(review_groups, OUTPUT_DIR / "NO_MATCH_NAMES.md")
    write_review_pdf(review_groups, OUTPUT_DIR / "NO_MATCH_NAMES.pdf")
    print(
        "Captain mailout manifest "
        f"mode={args.captain_attachment_mode} jobs={len(all_captain_jobs)}"
    )


if __name__ == "__main__":
    main()
